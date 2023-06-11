from typing import Iterable

from openpyxl.styles import Alignment
from openpyxl import Workbook


class TableBuilder:
    def __init__(self, sheet_name: str = None):
        self.workbook = Workbook()

        if sheet_name:
            self.workbook.remove(self.workbook.active)
            self.set_page(sheet_name)
        else:
            self.worksheet = self.workbook.active

    def set_page(self, name: str):
        if name in self.workbook.sheetnames:
            self.worksheet = self.workbook[name]
        else:
            self.worksheet = self.workbook.create_sheet(name)

    def add_line(self, data: Iterable, text_wrap_index: list[int] = None):
        self.worksheet.append(data)
        if text_wrap_index:
            for index in text_wrap_index:
                cell = self.worksheet._cells.get((self.worksheet._current_row, index), None)
                if cell:
                    cell.alignment = Alignment(wrapText=True)

    def write(self, file):
        self.workbook.save(file)
